"""
FastAPI 应用入口。

职责：
- 定义所有 API 路由（价格查询、状态、手动刷新、Excel 导出、历史记录）
- 管理应用生命周期（启动时初始化数据库连接池、执行首次价格抓取；关闭时释放资源）
- 通过 APScheduler 调度定时任务（每小时刷新价格、每日导出 CSV、每日清理过期历史）
- 挂载前端静态文件，并将 index.html 作为兜底路由
"""

import asyncio
import io
import logging
import os
import traceback
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone

from fastapi import FastAPI, Query, Header, HTTPException, Depends
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from fastapi.responses import StreamingResponse
from typing import Optional
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from openpyxl import Workbook

from pricing_data import (
    get_history, get_history_by_slug, get_prices, get_status,
    update_prices, update_yunwu_prices, update_litellm_prices, update_moyu_prices, PRICING_DATA,
)
import pricing_data
from price_fetcher import fetch_prices
from yunwu_fetcher import fetch_yunwu_prices
from litellm_fetcher import fetch_litellm_prices
from moyu_fetcher import fetch_moyu_prices
from csv_exporter import export_daily_csv, cleanup_old_csv
from database import init_pool, close_pool, insert_history_batch, cleanup_old_history

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
LOG_DIR = os.path.join(os.path.dirname(__file__), "logs")

scheduler = AsyncIOScheduler()
CHINA_TZ = timezone(timedelta(hours=8))

# 触发来源标签
_SOURCE_SCHEDULED = "SCHEDULED"   # 定时刷新
_SOURCE_MANUAL = "MANUAL"         # 用户通过网页 Refresh 按钮手动触发
_SOURCE_STARTUP = "STARTUP"       # 应用启动时首次加载


def _write_refresh_log(success: bool, model_count: int = 0, error_msg: str = "",
                       source: str = _SOURCE_SCHEDULED, platform: str = "openrouter",
                       exception: Exception | None = None):
    """将刷新结果写入 refresh.log；失败时额外生成独立的错误日志文件。"""
    log_dir = LOG_DIR
    os.makedirs(log_dir, exist_ok=True)
    now = datetime.now(CHINA_TZ)
    timestamp = now.strftime("%Y-%m-%d %H:%M:%S CST")

    # 追加到主刷新日志
    log_path = os.path.join(log_dir, "refresh.log")
    tag = f"[{source}]" if platform == "openrouter" else f"[{source}:{platform}]"
    if success:
        line = f"[{timestamp}] {tag} SUCCESS — refreshed {model_count} models\n"
    else:
        line = f"[{timestamp}] {tag} FAILED — {error_msg}\n"
    with open(log_path, "a") as f:
        f.write(line)

    # 失败时额外生成独立的错误日志文件
    if not success:
        error_filename = f"error_{now.strftime('%Y%m%d_%H%M%S')}_{platform}.log"
        error_path = os.path.join(log_dir, error_filename)
        with open(error_path, "w") as f:
            f.write(f"Platform:  {platform}\n")
            f.write(f"Source:    {source}\n")
            f.write(f"Time:      {timestamp}\n")
            f.write(f"Error:     {error_msg}\n")
            if exception:
                f.write(f"\nException type: {type(exception).__module__}.{type(exception).__qualname__}\n")
                f.write(f"Exception args: {exception.args}\n")
                f.write(f"\nTraceback:\n{''.join(traceback.format_exception(exception))}\n")
                if hasattr(exception, '__cause__') and exception.__cause__:
                    f.write(f"Caused by: {exception.__cause__}\n")


async def refresh_prices():
    """从 OpenRouter 拉取最新价格，更新内存数据并写入数据库历史。"""
    models = await fetch_prices()
    if models:
        try:
            update_prices(models)
        except Exception as e:
            logger.exception("update_prices failed: %s", e)
            raise RuntimeError(f"update_prices failed: {e}") from e
        now = datetime.now(timezone.utc)
        await insert_history_batch("openrouter", models, now)
        logger.info("Prices updated: %d models", len(models))
        return True, len(models)
    else:
        logger.warning("Refresh returned no data; keeping previous prices")
        now = datetime.now(timezone.utc)
        fallback = pricing_data._live_data if pricing_data._live_data else PRICING_DATA
        await insert_history_batch("openrouter", fallback, now)
        logger.info("Recorded history from current data (fallback)")
        return False, 0


async def refresh_yunwu_prices():
    """从云雾 AI 平台拉取最新价格，更新内存数据并写入数据库历史。"""
    models = await fetch_yunwu_prices()
    if models:
        try:
            update_yunwu_prices(models)
        except Exception as e:
            logger.exception("update_yunwu_prices failed: %s", e)
            raise RuntimeError(f"update_yunwu_prices failed: {e}") from e
        now = datetime.now(timezone.utc)
        await insert_history_batch("yunwu", models, now)
        logger.info("Yunwu prices updated: %d models", len(models))
        return True, len(models)
    else:
        logger.warning("Yunwu refresh returned no data")
        return False, 0


async def refresh_litellm_prices():
    """从 LiteLLM 定价数据库拉取官方厂商定价，更新内存数据并写入数据库历史。"""
    models = await fetch_litellm_prices()
    if models:
        try:
            update_litellm_prices(models)
        except Exception as e:
            logger.exception("update_litellm_prices failed: %s", e)
            raise RuntimeError(f"update_litellm_prices failed: {e}") from e
        now = datetime.now(timezone.utc)
        await insert_history_batch("official", models, now)
        logger.info("LiteLLM prices updated: %d models", len(models))
        return True, len(models)
    else:
        logger.warning("LiteLLM refresh returned no data")
        return False, 0


async def refresh_moyu_prices():
    """从魔芋平台拉取最新价格，更新内存数据并写入数据库历史。"""
    models = await fetch_moyu_prices()
    if models:
        try:
            update_moyu_prices(models)
        except Exception as e:
            logger.exception("update_moyu_prices failed: %s", e)
            raise RuntimeError(f"update_moyu_prices failed: {e}") from e
        now = datetime.now(timezone.utc)
        await insert_history_batch("moyu", models, now)
        logger.info("Moyu prices updated: %d models", len(models))
        return True, len(models)
    else:
        logger.warning("Moyu refresh returned no data")
        return False, 0


async def refresh_all(source: str = _SOURCE_SCHEDULED):
    """并发刷新所有平台的价格，将结果写入日志。"""
    results = await asyncio.gather(
        refresh_prices(),
        refresh_yunwu_prices(),
        refresh_litellm_prices(),
        refresh_moyu_prices(),
        return_exceptions=True,
    )

    or_result = results[0]
    yw_result = results[1]
    lt_result = results[2]
    my_result = results[3]

    if isinstance(or_result, Exception):
        _write_refresh_log(False, 0, str(or_result), source=source, platform="openrouter",
                           exception=or_result)
    else:
        or_ok, or_count = or_result
        _write_refresh_log(or_ok, or_count,
                           "No data returned from OpenRouter" if not or_ok else "",
                           source=source, platform="openrouter")

    if isinstance(yw_result, Exception):
        _write_refresh_log(False, 0, str(yw_result), source=source, platform="yunwu",
                           exception=yw_result)
    else:
        yw_ok, yw_count = yw_result
        _write_refresh_log(yw_ok, yw_count,
                           "No data returned from Yunwu" if not yw_ok else "",
                           source=source, platform="yunwu")

    if isinstance(lt_result, Exception):
        _write_refresh_log(False, 0, str(lt_result), source=source, platform="official",
                           exception=lt_result)
    else:
        lt_ok, lt_count = lt_result
        _write_refresh_log(lt_ok, lt_count,
                           "No data returned from LiteLLM" if not lt_ok else "",
                           source=source, platform="official")

    if isinstance(my_result, Exception):
        _write_refresh_log(False, 0, str(my_result), source=source, platform="moyu",
                           exception=my_result)
    else:
        my_ok, my_count = my_result
        _write_refresh_log(my_ok, my_count,
                           "No data returned from Moyu" if not my_ok else "",
                           source=source, platform="moyu")


async def scheduled_refresh():
    """定时刷新入口，捕获异常确保调度器不会因单次失败而终止。"""
    try:
        await refresh_all(source=_SOURCE_SCHEDULED)
    except Exception as e:
        logger.exception("Scheduled refresh crashed: %s", e)
        try:
            _write_refresh_log(False, 0, f"Exception: {e}", source=_SOURCE_SCHEDULED)
        except Exception:
            logger.exception("Failed to write refresh log after crash")


def daily_csv_task():
    """每日 CSV 快照任务：导出当日数据并清理过期文件。"""
    export_daily_csv(DATA_DIR)
    cleanup_old_csv(DATA_DIR)

def arithmetic_sequence(start, end, step=1):
    """生成等差数列字符串，用于 cron 表达式（如 '1,2,3,...,14'）。"""
    return ','.join(str(i) for i in range(start, end + 1, step))


@asynccontextmanager
async def lifespan(app: FastAPI):
    """应用生命周期管理：启动时初始化数据库、首次抓取、注册定时任务；关闭时释放资源。"""
    os.makedirs(DATA_DIR, exist_ok=True)

    db_ok = await init_pool()
    if not db_ok:
        logger.warning("Starting without database — history unavailable")

    # 启动时立即执行一次全平台价格刷新
    await refresh_all(source=_SOURCE_STARTUP)
    daily_csv_task()

    # 定时任务：每日中国时间 09:00–22:00 每小时整点刷新价格（UTC 01:00–14:00）
    scheduler.add_job(
        scheduled_refresh,
        "cron",
        hour=arithmetic_sequence(1, 13, 4),  # UTC 01:00, 05:00, 09:00, 13:00
        id="price_refresh",
        misfire_grace_time=300,
    )
    # 定时任务：每日 00:05 导出 CSV 快照
    scheduler.add_job(daily_csv_task, "cron", hour=0, minute=5, id="daily_csv",
                      misfire_grace_time=300)
    # 定时任务：每日 UTC 03:00 清理超过 7 天的历史数据
    scheduler.add_job(cleanup_old_history, "cron", hour=3, minute=0,
                      id="history_cleanup", misfire_grace_time=300)
    scheduler.start()
    logger.info("Scheduler started (price refresh hourly 09:00–22:00 CST, daily CSV at 00:05, history cleanup at 03:00 UTC)")
    yield
    scheduler.shutdown()
    await close_pool()


app = FastAPI(title="LLM Pricing Dashboard", lifespan=lifespan)
app.mount("/static", StaticFiles(directory="/app/static"), name="static")


# ————— API 路由 —————

@app.get("/api/prices")
def prices(provider: Optional[str] = Query(None), platform: Optional[str] = Query(None),
           model_type: Optional[str] = Query(None), model: Optional[str] = Query(None)):
    """获取当前所有模型定价，支持按供应商、平台、模型类型和模型名筛选。"""
    return get_prices(provider, platform, model_type, model)


@app.get("/api/status")
def status():
    """获取数据源状态（来源、更新时间、模型数量）。"""
    return get_status()


@app.post("/api/refresh")
async def manual_refresh():
    """手动触发全平台价格刷新。"""
    await refresh_all(source=_SOURCE_MANUAL)
    return get_status()


@app.get("/api/export")
def export_excel(provider: Optional[str] = Query(None), platform: Optional[str] = Query(None),
                 model_type: Optional[str] = Query(None), model: Optional[str] = Query(None)):
    """导出当前定价数据为 Excel (.xlsx) 文件并返回下载流。"""
    rows = get_prices(provider, platform, model_type, model)
    wb = Workbook()
    ws = wb.active
    ws.title = "LLM Pricing"
    headers = ["Platform", "Provider", "Model", "Type", "Input $/1M tokens", "Output $/1M tokens", "Context Window"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
    for r in rows:
        ctx = r["context_window"] if r["context_window"] else ""
        ws.append([r["platform"], r["provider"], r["model"], r.get("model_type", "text"),
                   r["input_price"], r["output_price"], ctx])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=llm_pricing.xlsx"},
    )


@app.get("/api/history")
async def history(provider: str = Query(...), model: str = Query(...),
            platform: Optional[str] = Query(None)):
    """查询指定模型的价格历史（按供应商+模型名）。"""
    return await get_history(provider, model, platform)


@app.get("/api/history/{slug}")
async def history_by_slug(slug: str):
    """按 URL slug 查询模型价格历史。"""
    result = await get_history_by_slug(slug)
    if not result:
        return JSONResponse(status_code=404, content={"detail": "Unknown model slug"})
    return result


# ————— 外部 API（需要 API Key 认证） —————

_API_KEY = os.environ.get("LLM_PRICING_API_KEY", "")


async def _verify_api_key(x_api_key: str = Header(None)):
    if not _API_KEY:
        raise HTTPException(status_code=503, detail="External API not configured (no API key set)")
    if x_api_key != _API_KEY:
        raise HTTPException(status_code=401, detail="Invalid or missing API key")


@app.get("/api/external/prices", dependencies=[Depends(_verify_api_key)])
def external_prices(provider: Optional[str] = Query(None), platform: Optional[str] = Query(None),
                    model_type: Optional[str] = Query(None), model: Optional[str] = Query(None)):
    """外部接口：获取当前所有模型定价（需 X-API-Key 请求头）。"""
    data = get_prices(provider, platform, model_type, model)
    status = get_status()
    return {
        "status": status,
        "count": len(data),
        "prices": data,
    }


@app.get("/api/external/status", dependencies=[Depends(_verify_api_key)])
def external_status():
    """外部接口：获取数据源状态（需 X-API-Key 请求头）。"""
    return get_status()


# ————— 前端页面路由 —————

@app.get("/history")
def history_page():
    """返回历史价格页面。"""
    return FileResponse("/app/static/history.html")


@app.get("/{slug}")
def history_slug_page(slug: str):
    """按 slug 路径访问历史价格页面（如 /gpt-4o）。"""
    return FileResponse("/app/static/history.html")


@app.get("/{full_path:path}")
def serve_frontend(full_path: str):
    """兜底路由：所有未匹配的路径返回主页 index.html。"""
    return FileResponse("/app/static/index.html")
