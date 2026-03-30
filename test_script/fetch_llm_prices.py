"""
Fetch LLM model prices from OpenRouter API and export to Excel.

Usage:
    python fetch_llm_prices.py

Output:
    llm_prices_YYYYMMDD_HHMMSS.xlsx in the same directory as this script.
"""

import os
import sys
from datetime import datetime, timezone

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

OPENROUTER_API = "https://openrouter.ai/api/v1/models"

PROVIDER_MAP = {
    "openai": "OpenAI",
    "anthropic": "Anthropic",
    "google": "Google",
    "deepseek": "DeepSeek",
    "mistralai": "Mistral",
    "mistral": "Mistral",
    "meta-llama": "Meta",
    "meta": "Meta",
}

# Static fallback data (used when the API is unreachable)
FALLBACK_DATA = [
    {"provider": "OpenAI", "model": "gpt-4o", "input_price": 2.50, "output_price": 10.00, "context_window": 128000},
    {"provider": "OpenAI", "model": "gpt-4o-mini", "input_price": 0.15, "output_price": 0.60, "context_window": 128000},
    {"provider": "OpenAI", "model": "gpt-4-turbo", "input_price": 10.00, "output_price": 30.00, "context_window": 128000},
    {"provider": "OpenAI", "model": "o1", "input_price": 15.00, "output_price": 60.00, "context_window": 200000},
    {"provider": "OpenAI", "model": "o1-mini", "input_price": 3.00, "output_price": 12.00, "context_window": 128000},
    {"provider": "OpenAI", "model": "o3-mini", "input_price": 1.10, "output_price": 4.40, "context_window": 200000},
    {"provider": "Anthropic", "model": "claude-opus-4", "input_price": 15.00, "output_price": 75.00, "context_window": 200000},
    {"provider": "Anthropic", "model": "claude-sonnet-4", "input_price": 3.00, "output_price": 15.00, "context_window": 200000},
    {"provider": "Anthropic", "model": "claude-haiku-3.5", "input_price": 0.80, "output_price": 4.00, "context_window": 200000},
    {"provider": "DeepSeek", "model": "deepseek-v3", "input_price": 0.27, "output_price": 1.10, "context_window": 128000},
    {"provider": "DeepSeek", "model": "deepseek-r1", "input_price": 0.55, "output_price": 2.19, "context_window": 128000},
    {"provider": "Google", "model": "gemini-2.0-flash", "input_price": 0.10, "output_price": 0.40, "context_window": 1000000},
    {"provider": "Google", "model": "gemini-2.0-pro", "input_price": 1.25, "output_price": 10.00, "context_window": 2000000},
    {"provider": "Google", "model": "gemini-1.5-pro", "input_price": 1.25, "output_price": 5.00, "context_window": 2000000},
    {"provider": "Mistral", "model": "mistral-large", "input_price": 2.00, "output_price": 6.00, "context_window": 128000},
    {"provider": "Mistral", "model": "mistral-small", "input_price": 0.10, "output_price": 0.30, "context_window": 128000},
    {"provider": "Mistral", "model": "codestral", "input_price": 0.30, "output_price": 0.90, "context_window": 256000},
    {"provider": "Meta", "model": "llama-3.3-70b", "input_price": 0.59, "output_price": 0.79, "context_window": 128000},
    {"provider": "Meta", "model": "llama-3.1-405b", "input_price": 3.00, "output_price": 3.00, "context_window": 128000},
    {"provider": "Meta", "model": "llama-3.1-8b", "input_price": 0.05, "output_price": 0.08, "context_window": 128000},
]

# ---------------------------------------------------------------------------
# Fetch prices
# ---------------------------------------------------------------------------


def parse_model(entry: dict) -> dict | None:
    """Parse a single model entry from the OpenRouter API response."""
    model_id = entry.get("id", "")
    parts = model_id.split("/", 1)
    if len(parts) != 2:
        return None

    slug, model_name = parts

    # Skip free/nitro/floor variants
    if ":" in model_name:
        return None

    provider = PROVIDER_MAP.get(slug.lower())
    if not provider:
        return None

    pricing = entry.get("pricing", {})
    prompt_price = pricing.get("prompt")
    completion_price = pricing.get("completion")
    if prompt_price is None or completion_price is None:
        return None

    try:
        input_per_million = float(prompt_price) * 1_000_000
        output_per_million = float(completion_price) * 1_000_000
    except (ValueError, TypeError):
        return None

    context_window = entry.get("context_length", 0)
    if not context_window:
        return None

    return {
        "provider": provider,
        "model": model_name,
        "input_price": round(input_per_million, 2),
        "output_price": round(output_per_million, 2),
        "context_window": context_window,
    }


def fetch_prices() -> tuple[list[dict], str]:
    """
    Fetch live prices from OpenRouter.
    Returns (models_list, source) where source is 'live' or 'static'.
    """
    try:
        print("Fetching prices from OpenRouter API...")
        resp = requests.get(OPENROUTER_API, timeout=30)
        resp.raise_for_status()
        data = resp.json().get("data", [])

        models = []
        seen = set()
        for entry in data:
            parsed = parse_model(entry)
            if parsed:
                key = (parsed["provider"], parsed["model"])
                if key not in seen:
                    seen.add(key)
                    models.append(parsed)

        if models:
            print(f"  Fetched {len(models)} models from OpenRouter (live data).")
            return models, "live"
        else:
            print("  OpenRouter returned no usable models, using fallback data.")
            return FALLBACK_DATA, "static"

    except Exception as e:
        print(f"  Failed to fetch from OpenRouter: {e}")
        print("  Using static fallback data.")
        return FALLBACK_DATA, "static"


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
EVEN_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

COLUMNS = [
    ("Provider", 14),
    ("Model", 28),
    ("Input Price ($/1M tokens)", 26),
    ("Output Price ($/1M tokens)", 27),
    ("Context Window (tokens)", 25),
]


def write_excel(models: list[dict], source: str, output_path: str) -> None:
    """Write model pricing data to an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "LLM Prices"

    # -- Title row --
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    title_cell.value = f"LLM Model Pricing  —  Source: {source}  |  Generated: {now_str}"
    title_cell.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # -- Header row (row 3) --
    header_row = 3
    for col_idx, (header_text, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # -- Sort: by provider then model name --
    sorted_models = sorted(models, key=lambda m: (m["provider"], m["model"]))

    # -- Data rows --
    for i, m in enumerate(sorted_models):
        row = header_row + 1 + i
        values = [
            m["provider"],
            m["model"],
            m["input_price"],
            m["output_price"],
            m["context_window"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if i % 2 == 0:
                cell.fill = EVEN_ROW_FILL
            # Number formatting
            if col_idx in (3, 4):
                cell.number_format = '#,##0.00'
            elif col_idx == 5:
                cell.number_format = '#,##0'

    # -- Auto-filter --
    last_row = header_row + len(sorted_models)
    ws.auto_filter.ref = f"A{header_row}:E{last_row}"

    # -- Freeze header --
    ws.freeze_panes = f"A{header_row + 1}"

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    models, source = fetch_prices()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_file = os.path.join(script_dir, f"llm_prices_{timestamp}.xlsx")

    write_excel(models, source, output_file)
    print(f"\nExcel file saved: {output_file}")
    print(f"  Total models: {len(models)}")


if __name__ == "__main__":
    main()
